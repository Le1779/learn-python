import openpyxl
from openpyxl.styles import Font 
import os

wb = openpyxl.load_workbook('test.xlsx')
count = 1
for sheet in wb.worksheets:
    if count == 4:
        continue
    rows = sheet.max_row
    for i in range(1, rows + 1):
        cell = sheet.cell(row = i, column = 2)
        if cell.value == None:
            continue

        cell.font = Font(name=cell.value)
    count = count + 1

wb.save('result.xlsx')